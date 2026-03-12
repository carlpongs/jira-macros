"""Generate a sample multi-issue Jira export in the example.xltm format for testing."""
import openpyxl
import random
from datetime import datetime, timedelta

CATEGORIES = [
    "Exercise Planning", "ASW Vignette", "Lethality Vignette", "Resupply Vignette",
    "Military Engagement", "DVPlanning", "C2", "Spectrum/Comms", "Networks",
    "Aerial Operations and Platforms", "Surface Operations and Platforms",
    "Undersea Operations and Platforms", "Afloat Logistics", "Mainland Logistics",
    "SCI Logistics", "Planning Conferences", "Security", "Data Collection",
    "Project Management", "Everything Else"
]

ASSIGNEES = [
    "Hannah G. Brood (NAVWAR)", "John D. Smith (PEO C4I)", "Sarah L. Jones (NIWC)",
    "Mike R. Chen (SPAWAR)", "Emily K. Davis (NAVSEA)", "Robert P. Wilson (ONR)",
    "Cmdr. A. Patel (PACFLT)", "Lt. B. Jackson (SURFPAC)"
]

STATUSES = ["To Do", "In Progress", "Done", "In Review", "Blocked"]
PRIORITIES = ["Normal", "High", "Critical", "Low"]
COMPONENTS = ["AUKUS MBP", "RIMPAC 2026", "PMINT", "C2 Integration", "ISR Systems"]

SUMMARIES = [
    "Determine designation as a Loitering Munition v. UAS and implications",
    "Review frequency allocation plan for joint operations",
    "Coordinate with allied forces on resupply chain logistics",
    "Update network architecture diagrams for exercise area",
    "Assess undersea sensor placement for ASW operations",
    "Draft DV engagement schedule for visiting officials",
    "Establish C2 node redundancy plan",
    "Validate spectrum deconfliction procedures",
    "Test satellite communication backup systems",
    "Plan aerial ISR coverage for exercise area",
    "Coordinate surface vessel movement plans",
    "Review security clearance requirements for participants",
    "Prepare data collection methodology document",
    "Schedule planning conference with coalition partners",
    "Assess mainland logistics support requirements",
    "Develop afloat logistics replenishment schedule",
    "Update SCI handling procedures for exercise",
    "Review project timeline and milestone tracking",
    "Conduct risk assessment for exercise scenarios",
    "Finalize participant roster and role assignments",
    "Evaluate communications interoperability gaps",
    "Draft after-action report template",
    "Coordinate with host nation for port access",
    "Review rules of engagement for exercise vignettes",
    "Prepare weather contingency plans",
    "Test tactical data link configurations",
    "Assess medical evacuation procedures",
    "Plan logistics for forward operating base",
    "Review intelligence sharing agreements",
    "Coordinate UAV flight path deconfliction",
]

DESCRIPTIONS = [
    "Does designation of how we can approach this matter",
    "Need to review current procedures and update accordingly",
    "Coordination required across multiple commands",
    "Impact assessment pending from technical review team",
    "Follow up with stakeholders for final determination",
    "Requires input from coalition partners before proceeding",
    "Timeline dependent on external approvals",
    "Budget implications need to be assessed",
]

COMMENTS = [
    ("Hannah G.", "Spoke with NAVWAR rep, awaiting response"),
    ("John D.", "Updated the requirements document"),
    ("Sarah L.", "Technical review completed, minor issues found"),
    ("Mike R.", "Coordinating with PEO C4I on schedule"),
    ("Emily K.", "Draft submitted for approval"),
    ("Robert P.", "Need additional funding authorization"),
]


def generate_sample(output_path, num_issues=45):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    issue_num = 1
    row = 1

    for _ in range(num_issues):
        cat = random.choice(CATEGORIES)
        status = random.choices(STATUSES, weights=[25, 35, 20, 10, 10])[0]
        created = datetime.now() - timedelta(days=random.randint(5, 60))
        updated = created + timedelta(days=random.randint(0, 10))

        # Issue header
        ws.cell(row=row, column=1, value=f"Issue CAT-{issue_num}")
        row += 1

        # Summary
        ws.cell(row=row, column=1, value="Summary:")
        ws.cell(row=row, column=2, value=random.choice(SUMMARIES))
        row += 1

        # Reporter / Issue Type
        ws.cell(row=row, column=1, value="Reporter:")
        ws.cell(row=row, column=2, value=random.choice(ASSIGNEES).split(" (")[0])
        ws.cell(row=row, column=6, value="Issue Type:")
        ws.cell(row=row, column=8, value="Task")
        row += 1

        # Assignee / Priority
        ws.cell(row=row, column=1, value="Assignee:")
        ws.cell(row=row, column=2, value=random.choice(ASSIGNEES))
        ws.cell(row=row, column=6, value="Priority:")
        ws.cell(row=row, column=8, value=random.choice(PRIORITIES))
        row += 1

        # URL
        ws.cell(row=row, column=1, value="URL:")
        ws.cell(row=row, column=2, value=f"CAT-{issue_num}")
        row += 1

        # Details header
        ws.cell(row=row, column=1, value="Details")
        row += 1

        # Description
        ws.cell(row=row, column=1, value="Description:")
        ws.cell(row=row, column=2, value=random.choice(DESCRIPTIONS))
        row += 1

        # Status
        ws.cell(row=row, column=1, value="Status:")
        ws.cell(row=row, column=2, value=status)
        row += 1

        # Resolution / Created
        ws.cell(row=row, column=1, value="Resolution:")
        ws.cell(row=row, column=2, value="Unresolved" if status != "Done" else "Done")
        ws.cell(row=row, column=6, value="Created:")
        ws.cell(row=row, column=8, value=created.strftime("%Y-%m-%d"))
        row += 1

        # Affects Version / Updated
        ws.cell(row=row, column=1, value="Affects Version/s:")
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=6, value="Updated:")
        ws.cell(row=row, column=8, value=updated.strftime("%Y-%m-%d"))
        row += 1

        # Fix Version
        ws.cell(row=row, column=1, value="Fix Version/s:")
        ws.cell(row=row, column=2, value="")
        row += 1

        # Component / Original Estimate
        ws.cell(row=row, column=1, value="Component/s:")
        ws.cell(row=row, column=2, value=random.choice(COMPONENTS))
        ws.cell(row=row, column=6, value="Original Estimate (Days):")
        ws.cell(row=row, column=8, value=random.randint(0, 10))
        row += 1

        # Labels / Time Spent
        ws.cell(row=row, column=1, value="Labels:")
        ws.cell(row=row, column=2, value=cat)
        ws.cell(row=row, column=6, value="Time Spent (Days):")
        ws.cell(row=row, column=8, value=random.randint(0, 5))
        row += 1

        # Sub-Tasks header
        ws.cell(row=row, column=1, value="Sub-Tasks")
        row += 1
        ws.cell(row=row, column=1, value="Key")
        ws.cell(row=row, column=2, value="Summary")
        ws.cell(row=row, column=7, value="Status")
        ws.cell(row=row, column=8, value="Assignee")
        row += 1
        ws.cell(row=row, column=8, value="Totals:")
        ws.cell(row=row, column=9, value=0)
        row += 1

        # Issue Links header
        ws.cell(row=row, column=1, value="Issue Links")
        row += 1
        ws.cell(row=row, column=1, value="Link Type")
        ws.cell(row=row, column=2, value="Issue Type")
        ws.cell(row=row, column=3, value="Key")
        ws.cell(row=row, column=4, value="Summary")
        ws.cell(row=row, column=8, value="Priority")
        ws.cell(row=row, column=9, value="Status")
        row += 1
        ws.cell(row=row, column=8, value="Totals:")
        ws.cell(row=row, column=9, value=0)
        row += 1

        # Comments header
        ws.cell(row=row, column=1, value="Comments")
        row += 1
        ws.cell(row=row, column=1, value="Author")
        ws.cell(row=row, column=2, value="Date")
        ws.cell(row=row, column=4, value="Body")
        row += 1

        # Add 0-2 comments
        num_comments = random.randint(0, 2)
        for _ in range(num_comments):
            author, body = random.choice(COMMENTS)
            comment_date = updated - timedelta(days=random.randint(0, 5))
            ws.cell(row=row, column=1, value=author)
            ws.cell(row=row, column=2, value=comment_date.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=4, value=body)
            row += 1

        ws.cell(row=row, column=8, value="Totals:")
        ws.cell(row=row, column=9, value=num_comments)
        row += 1

        # Worklogs header
        ws.cell(row=row, column=1, value="Worklogs")
        row += 1
        ws.cell(row=row, column=1, value="Author")
        ws.cell(row=row, column=2, value="Start Date")
        ws.cell(row=row, column=3, value="Time Spent")
        ws.cell(row=row, column=4, value="Comment")
        row += 1

        # Blank row separator
        row += 1
        issue_num += 1

    wb.save(output_path)
    print(f"Generated {num_issues} sample issues across {len(CATEGORIES)} categories -> {output_path}")


if __name__ == "__main__":
    import os as _os
    _script_dir = _os.path.dirname(_os.path.abspath(__file__))
    generate_sample(_os.path.join(_script_dir, "sample_jira_export.xlsx"), num_issues=45)


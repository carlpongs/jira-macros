#!/usr/bin/env python3
"""
Jira Kanban Export → Action Item Tracker
Converts a multi-issue Jira Excel export into a formatted one-page
Action Item Tracker Excel document, grouped by category.

Optimized for Windows — uses Segoe UI / Aptos fonts, os.startfile().
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import sys
import re
import collections
import subprocess

# ─── Category order from the Jira board ─────────────────────────────────────
CATEGORY_ORDER = [
    "Exercise Planning", "ASW Vignette", "Lethality Vignette",
    "Resupply Vignette", "Military Engagement", "DVPlanning", "C2",
    "Spectrum/Comms", "Networks", "Aerial Operations and Platforms",
    "Surface Operations and Platforms", "Undersea Operations and Platforms",
    "Afloat Logistics", "Mainland Logistics", "SCI Logistics",
    "Planning Conferences", "Security", "Data Collection",
    "Project Management", "Everything Else"
]

# ─── Colors ──────────────────────────────────────────────────────────────────
DARK_BG = "#1e1e2e"
DARK_FG = "#cdd6f4"
ACCENT = "#89b4fa"
ACCENT2 = "#a6e3a1"
SURFACE = "#313244"
OVERLAY = "#45475a"
RED = "#f38ba8"
YELLOW = "#f9e2af"
GREEN = "#a6e3a1"
PEACH = "#fab387"

# Excel colors (no #)
XL_HEADER_FILL = "1B2838"
XL_CAT_FILL = "2D4A6F"
XL_ROW_ALT1 = "F5F7FA"
XL_ROW_ALT2 = "FFFFFF"
XL_BORDER_COLOR = "B0B8C4"


# ═══════════════════════════════════════════════════════════════════════════════
#  PARSER — reads multi-block Jira Excel export
# ═══════════════════════════════════════════════════════════════════════════════
def parse_jira_export(filepath):
    """Parse a Jira Excel export in the example.xltm multi-row-per-issue format.
    Returns a list of dicts with keys: id, summary, assignee, status, labels,
    description, comments, component, priority, created, updated."""

    wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
    ws = wb.active

    issues = []
    current = None
    in_comments = False
    in_worklogs = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
        a_val = str(row[0].value or "").strip()
        b_val = str(row[1].value or "").strip() if len(row) > 1 and row[1].value else ""
        d_val = str(row[3].value or "").strip() if len(row) > 3 and row[3].value else ""
        h_val = str(row[7].value or "").strip() if len(row) > 7 and row[7].value else ""

        # New issue block
        if re.match(r"Issue\s+CAT-\d+", a_val):
            if current:
                issues.append(current)
            issue_id = a_val.replace("Issue ", "")
            current = {
                "id": issue_id, "summary": "", "assignee": "", "status": "",
                "labels": "", "description": "", "comments": [],
                "component": "", "priority": "", "created": "", "updated": ""
            }
            in_comments = False
            in_worklogs = False
            continue

        if not current:
            continue

        # Section headers
        if a_val == "Comments":
            in_comments = True
            in_worklogs = False
            continue
        elif a_val == "Worklogs":
            in_worklogs = True
            in_comments = False
            continue
        elif a_val in ("Sub-Tasks", "Issue Links", "Details"):
            in_comments = False
            in_worklogs = False
            continue

        # Skip table headers
        if a_val in ("Key", "Link Type", "Author") and not in_comments:
            continue

        # Parse comments
        if in_comments and a_val and a_val != "Author":
            if a_val != "" and b_val != "" and not a_val.startswith("Totals"):
                current["comments"].append({
                    "author": a_val, "date": b_val, "body": d_val
                })
            continue

        if in_worklogs:
            continue

        # Parse fields
        if a_val == "Summary:":
            current["summary"] = b_val
        elif a_val == "Assignee:":
            current["assignee"] = b_val
        elif a_val == "Status:":
            current["status"] = b_val
        elif a_val == "Labels:":
            current["labels"] = b_val
        elif a_val == "Description:":
            current["description"] = b_val
        elif a_val == "Component/s:":
            current["component"] = b_val

        # Fields in column F/H
        if a_val == "Reporter:" or a_val == "Assignee:":
            if h_val:
                if "Priority:" in str(row[5].value or ""):
                    current["priority"] = h_val
                elif "Issue Type:" in str(row[5].value or ""):
                    pass  # skip
        if a_val == "Resolution:":
            if h_val:
                current["created"] = h_val
        if a_val == "Affects Version/s:":
            if h_val:
                current["updated"] = h_val

    if current:
        issues.append(current)

    wb.close()
    return issues


def group_by_category(issues):
    """Group issues by their Labels field, ordered by CATEGORY_ORDER."""
    grouped = collections.defaultdict(list)
    for issue in issues:
        cat = issue.get("labels", "").strip()
        if not cat:
            cat = "Everything Else"
        grouped[cat].append(issue)

    # Return in defined order, then any extras
    ordered = []
    seen = set()
    for cat in CATEGORY_ORDER:
        if cat in grouped:
            ordered.append((cat, grouped[cat]))
            seen.add(cat)
    for cat in sorted(grouped.keys()):
        if cat not in seen:
            ordered.append((cat, grouped[cat]))
    return ordered


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL GENERATOR — Action Item Tracker output
# ═══════════════════════════════════════════════════════════════════════════════
def generate_action_item_tracker(issues, output_path):
    """Generate the formatted Action Item Tracker Excel file."""
    grouped = group_by_category(issues)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Action Item Tracker"

    # Page setup for one-page printing
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LETTER

    # Column widths  — 4 main columns; description & comments go in a detail row underneath
    NUM_COLS = 4
    ws.column_dimensions["A"].width = 12   # Issue ID  /  Description (left half)
    ws.column_dimensions["B"].width = 44   # Summary   /  Description (right half)
    ws.column_dimensions["C"].width = 22   # Assignee  /  Comments (left half)
    ws.column_dimensions["D"].width = 50   # Status    /  Comments (right half)

    # Styles
    thin_border = Border(
        left=Side(style="thin", color=XL_BORDER_COLOR),
        right=Side(style="thin", color=XL_BORDER_COLOR),
        top=Side(style="thin", color=XL_BORDER_COLOR),
        bottom=Side(style="thin", color=XL_BORDER_COLOR),
    )

    title_font = Font(name="Aptos", size=16, bold=True, color="1B2838")
    date_font = Font(name="Aptos", size=10, italic=True, color="555555")
    cat_font = Font(name="Aptos", size=12, bold=True, color="FFFFFF")
    cat_fill = PatternFill(start_color=XL_CAT_FILL, end_color=XL_CAT_FILL, fill_type="solid")
    header_font = Font(name="Aptos", size=9, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=XL_HEADER_FILL, end_color=XL_HEADER_FILL, fill_type="solid")
    data_font = Font(name="Aptos", size=9)
    desc_font = Font(name="Aptos", size=8, italic=True, color="444444")
    comment_font = Font(name="Aptos", size=8, color="666666")
    detail_label_font = Font(name="Aptos", size=8, bold=True, italic=True, color="333333")
    fill_alt1 = PatternFill(start_color=XL_ROW_ALT1, end_color=XL_ROW_ALT1, fill_type="solid")
    fill_alt2 = PatternFill(start_color=XL_ROW_ALT2, end_color=XL_ROW_ALT2, fill_type="solid")
    detail_fill = PatternFill(start_color="EDF0F5", end_color="EDF0F5", fill_type="solid")

    # Status colors
    status_colors = {
        "To Do": Font(name="Aptos", size=9, color="CC6600"),
        "In Progress": Font(name="Aptos", size=9, bold=True, color="0066CC"),
        "Done": Font(name="Aptos", size=9, color="228B22"),
        "In Review": Font(name="Aptos", size=9, color="8B008B"),
        "Blocked": Font(name="Aptos", size=9, bold=True, color="CC0000"),
    }

    row = 1

    # ── Title ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1, value="ACTION ITEM TRACKER")
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    # ── Date ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}")
    cell.font = date_font
    cell.alignment = Alignment(horizontal="center")
    row += 1

    # ── Stats summary ──
    total = len(issues)
    status_counts = collections.Counter(i["status"] for i in issues)
    stats = f"Total: {total}  |  "
    stats += "  |  ".join(f"{s}: {c}" for s, c in sorted(status_counts.items()))
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
    cell = ws.cell(row=row, column=1, value=stats)
    cell.font = Font(name="Aptos", size=9, bold=True, color="333333")
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # ── Categories ──
    for cat_name, cat_issues in grouped:
        # Category header — spans all 4 columns
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NUM_COLS)
        cell = ws.cell(row=row, column=1, value=f"▸ {cat_name} ({len(cat_issues)})")
        cell.font = cat_font
        cell.fill = cat_fill
        cell.alignment = Alignment(vertical="center")
        for col in range(1, NUM_COLS + 1):
            ws.cell(row=row, column=col).fill = cat_fill
            ws.cell(row=row, column=col).border = thin_border
        row += 1

        # Column headers — 4 columns
        headers = ["Issue ID", "Summary", "Assignee", "Status"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row += 1

        # Issue rows — each issue gets TWO rows:
        #   Row 1: Issue ID | Summary | Assignee | Status
        #   Row 2: Description (cols A-B)  |  Comments (cols C-D)
        for idx, issue in enumerate(cat_issues):
            fill = fill_alt1 if idx % 2 == 0 else fill_alt2

            # ── Main row ──
            values = [issue["id"], issue["summary"], issue["assignee"], issue["status"]]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.font = status_colors.get(val, data_font)
                cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=(col == 2))
            row += 1

            # ── Detail row (Description | Comments) ──
            desc_text = issue.get("description", "")
            comments_parts = []
            for c in issue.get("comments", []):
                comments_parts.append(f"{c['author']} ({c['date']}): {c['body']}")
            comments_text = "  |  ".join(comments_parts)

            # Only add detail row if there is something to show
            if desc_text or comments_text:
                # Description in merged cols A-B
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
                cell = ws.cell(row=row, column=1,
                               value=f"Desc: {desc_text}" if desc_text else "")
                cell.font = desc_font
                cell.fill = detail_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                # Apply fill/border to merged cell B
                ws.cell(row=row, column=2).fill = detail_fill
                ws.cell(row=row, column=2).border = thin_border

                # Comments in merged cols C-D
                ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
                cell = ws.cell(row=row, column=3,
                               value=f"Comments: {comments_text}" if comments_text else "")
                cell.font = comment_font
                cell.fill = detail_fill
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                # Apply fill/border to merged cell D
                ws.cell(row=row, column=4).fill = detail_fill
                ws.cell(row=row, column=4).border = thin_border
                row += 1

        # Spacer row
        row += 1

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════════════════════════
#  TKINTER GUI
# ═══════════════════════════════════════════════════════════════════════════════
class JiraReportApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Jira → Action Item Tracker")
        self.root.geometry("1000x650")
        self.root.configure(bg=DARK_BG)
        self.root.minsize(800, 500)

        self.issues = []
        self.filepath = None

        self._build_ui()

    def _build_ui(self):
        # ── Style ──
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Dark.TFrame", background=DARK_BG)
        style.configure("Dark.TLabel", background=DARK_BG, foreground=DARK_FG,
                         font=("Segoe UI", 10))
        style.configure("Title.TLabel", background=DARK_BG, foreground=ACCENT,
                         font=("Segoe UI", 18, "bold"))
        style.configure("Stat.TLabel", background=SURFACE, foreground=DARK_FG,
                         font=("Segoe UI", 11), padding=8)
        style.configure("Dark.TButton", background=SURFACE, foreground=DARK_FG,
                         font=("Segoe UI", 10, "bold"), padding=(12, 6))
        style.map("Dark.TButton",
                   background=[("active", OVERLAY)],
                   foreground=[("active", ACCENT)])
        style.configure("Accent.TButton", background=ACCENT, foreground=DARK_BG,
                         font=("Segoe UI", 10, "bold"), padding=(16, 8))
        style.map("Accent.TButton",
                   background=[("active", ACCENT2)])
        style.configure("Treeview", background=SURFACE, foreground=DARK_FG,
                         fieldbackground=SURFACE, font=("Segoe UI", 9),
                         rowheight=26)
        style.configure("Treeview.Heading", background=OVERLAY, foreground=ACCENT,
                         font=("Segoe UI", 9, "bold"))
        style.map("Treeview", background=[("selected", OVERLAY)])

        # ── Top bar ──
        top = ttk.Frame(self.root, style="Dark.TFrame")
        top.pack(fill="x", padx=16, pady=(16, 8))

        ttk.Label(top, text="⚡ Jira → Action Item Tracker", style="Title.TLabel").pack(side="left")

        btn_frame = ttk.Frame(top, style="Dark.TFrame")
        btn_frame.pack(side="right")

        self.btn_export = ttk.Button(btn_frame, text="📄 Export Tracker", style="Accent.TButton",
                                      command=self._export)
        self.btn_export.pack(side="right", padx=(8, 0))
        self.btn_export.state(["disabled"])

        ttk.Button(btn_frame, text="📂 Import Jira Export", style="Dark.TButton",
                   command=self._import).pack(side="right")

        # ── Stats bar ──
        self.stats_frame = ttk.Frame(self.root, style="Dark.TFrame")
        self.stats_frame.pack(fill="x", padx=16, pady=4)
        self.stat_labels = {}

        # ── Main area ──
        main = ttk.Frame(self.root, style="Dark.TFrame")
        main.pack(fill="both", expand=True, padx=16, pady=(4, 16))

        # Treeview
        cols = ("id", "summary", "assignee", "status", "category")
        self.tree = ttk.Treeview(main, columns=cols, show="headings", selectmode="browse")
        self.tree.heading("id", text="Issue ID")
        self.tree.heading("summary", text="Summary")
        self.tree.heading("assignee", text="Assignee")
        self.tree.heading("status", text="Status")
        self.tree.heading("category", text="Category")
        self.tree.column("id", width=80, minwidth=60)
        self.tree.column("summary", width=400, minwidth=200)
        self.tree.column("assignee", width=180, minwidth=100)
        self.tree.column("status", width=100, minwidth=70)
        self.tree.column("category", width=180, minwidth=100)

        scrollbar = ttk.Scrollbar(main, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # ── Status bar ──
        self.status_var = tk.StringVar(value="Ready — import a Jira Excel export to begin")
        status_bar = ttk.Label(self.root, textvariable=self.status_var, style="Dark.TLabel",
                               font=("Segoe UI", 9))
        status_bar.pack(fill="x", padx=16, pady=(0, 8))

    def _update_stats(self):
        # Clear old
        for w in self.stats_frame.winfo_children():
            w.destroy()

        if not self.issues:
            return

        total = len(self.issues)
        status_counts = collections.Counter(i["status"] for i in self.issues)
        cat_count = len(set(i["labels"] or "Everything Else" for i in self.issues))

        stats = [
            (f"📊 {total} Issues", ACCENT),
            (f"📁 {cat_count} Categories", PEACH),
        ]
        color_map = {"To Do": YELLOW, "In Progress": ACCENT, "Done": GREEN,
                     "Blocked": RED, "In Review": PEACH}
        for s, c in sorted(status_counts.items()):
            stats.append((f"{s}: {c}", color_map.get(s, DARK_FG)))

        for text, color in stats:
            lbl = tk.Label(self.stats_frame, text=text, bg=SURFACE, fg=color,
                           font=("Segoe UI", 10, "bold"), padx=12, pady=6)
            lbl.pack(side="left", padx=(0, 6))

    def _import(self):
        initial_dir = os.path.expanduser("~")
        path = filedialog.askopenfilename(
            title="Select Jira Export File",
            initialdir=initial_dir,
            filetypes=[("Excel files", "*.xlsx *.xltm *.xls"), ("All files", "*.*")]
        )
        if not path:
            return

        try:
            self.filepath = path
            self.issues = parse_jira_export(path)

            # Populate treeview
            self.tree.delete(*self.tree.get_children())
            for issue in self.issues:
                cat = issue.get("labels") or "Everything Else"
                self.tree.insert("", "end", values=(
                    issue["id"], issue["summary"][:80],
                    issue["assignee"], issue["status"], cat
                ))

            self._update_stats()
            self.btn_export.state(["!disabled"])
            self.status_var.set(f"✅ Loaded {len(self.issues)} issues from {os.path.basename(path)}")

        except Exception as e:
            messagebox.showerror("Import Error", f"Failed to parse file:\n{e}")
            self.status_var.set(f"❌ Error: {e}")

    def _export(self):
        if not self.issues:
            messagebox.showwarning("No Data", "Import a Jira export first.")
            return

        default_name = f"Action_Item_Tracker_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        initial_dir = os.path.dirname(self.filepath) if self.filepath else os.path.expanduser("~")
        path = filedialog.asksaveasfilename(
            title="Save Action Item Tracker",
            initialdir=initial_dir,
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not path:
            return

        try:
            generate_action_item_tracker(self.issues, path)
            self.status_var.set(f"✅ Exported to {os.path.basename(path)}")

            # Offer to open the file (Windows: os.startfile, others: xdg-open / open)
            open_it = messagebox.askyesno(
                "Export Complete",
                f"Action Item Tracker saved!\n\n{path}\n\nOpen the file now?"
            )
            if open_it:
                _open_file(path)

        except Exception as e:
            messagebox.showerror("Export Error", f"Failed to generate tracker:\n{e}")
            self.status_var.set(f"❌ Export error: {e}")


def _open_file(filepath):
    """Open a file with the system default application (cross-platform)."""
    try:
        if sys.platform == "win32":
            os.startfile(filepath)
        elif sys.platform == "darwin":
            subprocess.Popen(["open", filepath])
        else:
            subprocess.Popen(["xdg-open", filepath])
    except Exception:
        pass  # silently ignore if opener unavailable


def main():
    root = tk.Tk()
    app = JiraReportApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
